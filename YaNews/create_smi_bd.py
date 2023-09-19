from openpyxl import load_workbook
from parsing_smi_info import make_smi_profile
import time
import random


workbook = load_workbook(filename="smi_db.xlsm")
sheet = workbook['Main_type']

for value in sheet.iter_rows(min_col=2,
                             max_col=2,
                             min_row=7532,
                             # max_row=6883,
                             values_only=True):
    smi_link = value[0].strip()
    folder_name = smi_link.replace('https://news.yandex.ru/smi/', '')
    print(f'Записываю {smi_link}...')
    make_smi_profile(smi_link, folder_name)
    delay_time = random.randint(11, 15)
    time.sleep(delay_time)
