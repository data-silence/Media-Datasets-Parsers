import glob
from openpyxl import load_workbook
import shutil
# from parsing_smi_info import make_smi_profile


workbook = load_workbook(filename="smi_db.xlsm")
sheet = workbook['Main_type']
log = []

filenames_list = glob.glob("smi_bd\\*.txt")  # сохраняет в список файлы в директории articles с расширением *.txt

for value in sheet.iter_rows(min_col=2,
                             max_col=2,
                             min_row=2,
                             # max_row=6883,
                             values_only=True):
    try:
        smi_link = value[0].strip()
        folder_name = smi_link.replace('https://news.yandex.ru/smi/', '')
        if folder_name not in filenames_list:
            shutil.move(f'smi_bd\\{folder_name}\\smi_profile.txt', f'smi_bd\\{folder_name}.txt')
    except:
        log.append(folder_name)

with open('log.txt', 'w', encoding='utf8') as f:
    for el in log:
        f.write(f'{el}\n')
